"""Extract 2026 Trustees Report calibration targets.

Inputs (downloaded 2026-06-10; see data/tr2026_sources.manifest.json):
- SSA single-year TR tables workbook (``SingleYearTRTables_TR2026.xlsx``):
  - VI.G1: taxable payroll, GDP, average wage index (current dollars)
  - VI.G2: combined OASDI cost (current dollars)
  - IV.B2: OASDI income from taxation of benefits as % of taxable payroll
  - V.A3: Social Security area population by broad age group
- CMS 2026 Medicare Trustees expanded tables: ``Medicare Sources of
  Non-Interest Income …`` — HI income from taxation of benefits in
  millions of dollars, annual through 2100.

Outputs:
- ``data/social_security_aux_tr2026.csv`` — nominal OASDI cost, taxable
  payroll, GDP, AWI, OASDI TOB (% and $), HI TOB ($), 2025-2100.
- ``data/SSPopJul_TR2026_interim.csv`` — population by single year of age:
  the TR2024 single-year age shape rescaled within each TR2026 V.A3 age
  group (under 20 / 20-64 / 65+) so group totals match TR2026 exactly.
  Interim until SSA posts the TR2026 single-year-age file.
- ``data/tr2026_sources.manifest.json`` — source provenance and hashes.
"""

from __future__ import annotations

import csv
import hashlib
import json
from pathlib import Path

import openpyxl
import pandas as pd

REPO_ROOT = Path(__file__).resolve().parent.parent
SOURCES = REPO_ROOT / "data" / "sources" / "tr2026"
WORKBOOK = SOURCES / "SingleYearTRTables_TR2026.xlsx"
MEDICARE_CSV = (
    SOURCES / "Medicare Sources of Non-Interest Income as a Percentage of Total "
    "Income and as a Percentage of Gross Domestic Product.csv"
)
TR2024_POPULATION = REPO_ROOT / "data" / "SSPopJul_TR2024.csv"

AUX_OUTPUT = REPO_ROOT / "data" / "social_security_aux_tr2026.csv"
POPULATION_OUTPUT = REPO_ROOT / "data" / "SSPopJul_TR2026_interim.csv"
MANIFEST_OUTPUT = REPO_ROOT / "data" / "tr2026_sources.manifest.json"

HI_RATES_CSV = SOURCES / "HI Cost and Income Rates.csv"
TR2025_AUX = REPO_ROOT / "data" / "social_security_aux_tr2025.csv"
TR2025_HI = REPO_ROOT / "data" / "hi_expenditures_tr2025.csv"

ECONOMIC_PROJECTIONS_OUTPUTS = (
    REPO_ROOT / "data" / "ssa_economic_projections.csv",
    REPO_ROOT / "dashboard" / "public" / "data" / "ssa_economic_projections.csv",
)
HI_PAYROLL_OUTPUT = (
    REPO_ROOT / "dashboard" / "public" / "data" / "hi_taxable_payroll.csv"
)
TRUST_FUND_GAPS_OUTPUT = REPO_ROOT / "data" / "trust_fund_gaps.csv"

SOURCE_URLS = {
    "SingleYearTRTables_TR2026.xlsx": (
        "https://www.ssa.gov/oact/TR/2026/SingleYearTRTables_TR2026.xlsx"
        " (retrieved via web.archive.org, 2026-06-10)"
    ),
    MEDICARE_CSV.name: (
        "https://www.cms.gov/files/zip/"
        "2026-expanded-supplementary-tables-figures.zip (retrieved direct, "
        "2026-06-10)"
    ),
}

YEARS = range(2025, 2101)


def _sheet_year_series(worksheet, value_column: int, *, label: str) -> dict[int, float]:
    """Collect {year: value} from the Intermediate section of a
    single-year TR table sheet.

    The workbooks stack Historical, Intermediate, Low-cost, and High-cost
    sections; only the intermediate (II) projection is a calibration
    target.
    """
    series: dict[int, float] = {}
    section = "historical"
    for row in worksheet.iter_rows(values_only=True):
        first = row[0]
        if isinstance(first, str):
            marker = first.strip().lower()
            if marker.startswith("historical"):
                section = "historical"
            elif marker.startswith("intermediate"):
                section = "intermediate"
            elif marker.startswith(("low-cost", "high-cost")):
                section = "alternative"
            continue
        if section == "alternative":
            continue
        try:
            year = int(first)
        except (TypeError, ValueError):
            continue
        if not 1940 <= year <= 2100:
            continue
        value = row[value_column]
        if value is None or isinstance(value, str):
            continue
        if year in series and section != "intermediate":
            continue  # intermediate values win over historical estimates
        series[year] = float(value)
    if not series:
        raise ValueError(f"{label}: no year rows parsed")
    return series


def extract_ssa_series() -> pd.DataFrame:
    workbook = openpyxl.load_workbook(WORKBOOK, read_only=True)
    payroll = _sheet_year_series(workbook["VI.G1"], 3, label="VI.G1 taxable payroll")
    gdp = _sheet_year_series(workbook["VI.G1"], 4, label="VI.G1 GDP")
    awi = _sheet_year_series(workbook["VI.G1"], 2, label="VI.G1 AWI")
    # VI.G2 (cost in dollars) truncates at combined reserve depletion, so
    # cost comes from the IV.B1 OASDI cost rate (percent of payroll, which
    # runs through 2100) times taxable payroll.
    cost_rate = _sheet_year_series(workbook["IV.B1"], 8, label="IV.B1 OASDI cost rate")
    tob_pct = _sheet_year_series(
        workbook["IV.B2"], 10, label="IV.B2 OASDI TOB % of payroll"
    )

    rows = []
    for year in YEARS:
        rows.append(
            {
                "year": year,
                "oasdi_cost_in_billion_nominal_usd": round(
                    cost_rate[year] / 100.0 * payroll[year], 4
                ),
                "oasdi_cost_rate_pct_of_taxable_payroll": cost_rate[year],
                "taxable_payroll_in_billion_nominal_usd": payroll[year],
                "gdp_in_billion_nominal_usd": gdp[year],
                "average_wage_index": awi[year],
                "oasdi_tob_pct_of_taxable_payroll": tob_pct[year],
                "oasdi_tob_billions_nominal_usd": round(
                    tob_pct[year] / 100.0 * payroll[year], 4
                ),
            }
        )
    return pd.DataFrame(rows)


def extract_hi_tob() -> dict[int, float]:
    """HI income from taxation of benefits, billions of dollars by year."""
    with MEDICARE_CSV.open(encoding="utf-8-sig") as handle:
        rows = list(csv.reader(handle))
    header = next(row for row in rows if row and row[0].strip().startswith("Calendar"))
    tob_column = next(
        index
        for index, cell in enumerate(header)
        if "tax on" in cell.lower().replace("\n", " ")
    )
    series: dict[int, float] = {}
    for row in rows:
        cell = (row[0] if row else "").strip()
        if not cell.isdigit():
            continue
        year = int(cell)
        raw = row[tob_column].strip().replace(",", "").replace("-", "")
        if not raw:
            continue
        series[year] = float(raw) / 1_000.0  # millions -> billions
    missing = [year for year in YEARS if year not in series]
    if missing:
        raise ValueError(f"Medicare HI TOB missing years: {missing[:5]}")
    return series


def extract_population_groups() -> pd.DataFrame:
    """TR2026 V.A3 population by broad age group, in persons."""
    workbook = openpyxl.load_workbook(WORKBOOK, read_only=True)
    worksheet = workbook["V.A3"]
    under_20 = _sheet_year_series(worksheet, 1, label="V.A3 under 20")
    work_age = _sheet_year_series(worksheet, 2, label="V.A3 20-64")
    aged = _sheet_year_series(worksheet, 3, label="V.A3 65+")
    rows = [
        {
            "year": year,
            "under_20": under_20[year] * 1_000,
            "age_20_64": work_age[year] * 1_000,
            "age_65_plus": aged[year] * 1_000,
        }
        for year in YEARS
        if year in under_20
    ]
    return pd.DataFrame(rows)


def build_interim_single_year_population(groups: pd.DataFrame) -> pd.DataFrame:
    """Rescale the TR2024 single-year age shape to TR2026 group totals."""
    tr2024 = pd.read_csv(TR2024_POPULATION)[["Year", "Age", "Total"]]
    tr2024 = tr2024[tr2024.Year.isin(YEARS)]

    def group_of(age: pd.Series) -> pd.Series:
        return pd.cut(
            age,
            bins=[-1, 19, 64, 200],
            labels=["under_20", "age_20_64", "age_65_plus"],
        )

    tr2024 = tr2024.assign(group=group_of(tr2024.Age))
    group_2024 = tr2024.groupby(["Year", "group"], observed=True).Total.sum().unstack()
    targets = groups.set_index("year")

    frames = []
    for year in targets.index:
        if year not in group_2024.index:
            continue
        ratios = {
            group: targets.loc[year, group] / group_2024.loc[year, group]
            for group in ("under_20", "age_20_64", "age_65_plus")
        }
        year_frame = tr2024[tr2024.Year == year].copy()
        year_frame["Total"] = year_frame.Total * year_frame.group.map(ratios).astype(
            float
        )
        frames.append(year_frame[["Year", "Age", "Total"]])
    interim = pd.concat(frames, ignore_index=True)
    interim["Total"] = interim.Total.round(0).astype(int)
    return interim


def extract_hi_rates() -> pd.DataFrame:
    """HI cost and income rates (% of HI taxable payroll) by year."""
    with HI_RATES_CSV.open(encoding="utf-8-sig") as handle:
        rows = list(csv.reader(handle))
    records = []
    for row in rows:
        cell = (row[0] if row else "").strip()
        if not cell.isdigit():
            continue
        try:
            records.append(
                {
                    "year": int(cell),
                    "hi_cost_rate": float(row[1]),
                    "hi_income_rate": float(row[2]),
                }
            )
        except (ValueError, IndexError):
            continue
    frame = pd.DataFrame(records).drop_duplicates("year", keep="last")
    return frame.set_index("year")


def write_dashboard_denominators(aux: pd.DataFrame) -> None:
    """Refresh dashboard payroll/GDP denominators and trust-fund gaps."""
    projections = aux.rename(
        columns={
            "taxable_payroll_in_billion_nominal_usd": "taxable_payroll",
            "gdp_in_billion_nominal_usd": "gdp",
        }
    )[["year", "taxable_payroll", "gdp"]]
    projections = projections[projections.year >= 2026].round(1)
    for output in ECONOMIC_PROJECTIONS_OUTPUTS:
        projections.to_csv(output, index=False)
        print(f"wrote {output}")

    # HI taxable payroll: scale the TR2026 OASDI payroll by the TR2025
    # HI/OASDI payroll ratio (display denominator only; the Medicare
    # expanded tables do not publish HI payroll levels directly).
    tr2025 = pd.read_csv(TR2025_AUX).set_index("year")
    tr2025_hi = pd.read_csv(TR2025_HI)[["year", "hi_taxable_payroll"]]
    tr2025_hi["hi_taxable_payroll"] = tr2025_hi.hi_taxable_payroll / 1e9
    tr2025_hi = tr2025_hi.set_index("year")
    dashboard_hi = pd.read_csv(HI_PAYROLL_OUTPUT).set_index("year")
    rows = []
    for year in projections.year:
        if year in tr2025_hi.index:
            hi_2025 = float(tr2025_hi.loc[year, "hi_taxable_payroll"])
        elif year in dashboard_hi.index:
            hi_2025 = float(dashboard_hi.loc[year, "hi_taxable_payroll"])
        else:
            continue
        ratio = hi_2025 / float(
            tr2025.loc[year, "taxable_payroll_in_billion_nominal_usd"]
        )
        payroll_2026 = float(
            projections.loc[projections.year == year, "taxable_payroll"].iloc[0]
        )
        rows.append(
            {"year": year, "hi_taxable_payroll": round(ratio * payroll_2026, 6)}
        )
    pd.DataFrame(rows).to_csv(HI_PAYROLL_OUTPUT, index=False)
    print(f"wrote {HI_PAYROLL_OUTPUT} ({len(rows)} years)")

    workbook = openpyxl.load_workbook(WORKBOOK, read_only=True)
    oasdi_income = _sheet_year_series(
        workbook["IV.B1"], 7, label="IV.B1 OASDI income rate"
    )
    oasdi_cost = _sheet_year_series(workbook["IV.B1"], 8, label="IV.B1 OASDI cost rate")
    hi_rates = extract_hi_rates()
    gap_rows = []
    for year in YEARS:
        if year < 2026 or year not in hi_rates.index:
            continue
        gap_rows.append(
            {
                "year": year,
                "oasdi_cost_rate": round(oasdi_cost[year], 2),
                "oasdi_income_rate": round(oasdi_income[year], 2),
                "oasdi_gap_pct": round(oasdi_cost[year] - oasdi_income[year], 2),
                "hi_cost_rate": round(hi_rates.loc[year, "hi_cost_rate"], 2),
                "hi_income_rate": round(hi_rates.loc[year, "hi_income_rate"], 2),
                "hi_gap_pct": round(
                    hi_rates.loc[year, "hi_cost_rate"]
                    - hi_rates.loc[year, "hi_income_rate"],
                    2,
                ),
            }
        )
    pd.DataFrame(gap_rows).to_csv(TRUST_FUND_GAPS_OUTPUT, index=False)
    print(f"wrote {TRUST_FUND_GAPS_OUTPUT} ({len(gap_rows)} years)")


def file_sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def main() -> int:
    aux = extract_ssa_series()
    hi_tob = extract_hi_tob()
    aux["hi_tob_billions_nominal_usd"] = aux.year.map(hi_tob)
    aux.to_csv(AUX_OUTPUT, index=False)
    print(f"wrote {AUX_OUTPUT} ({len(aux)} years)")
    sample = aux[aux.year.isin([2026, 2050, 2100])].round(2)
    print(sample.to_string(index=False))

    write_dashboard_denominators(aux)

    groups = extract_population_groups()
    interim = build_interim_single_year_population(groups)
    interim.to_csv(POPULATION_OUTPUT, index=False)
    print(f"wrote {POPULATION_OUTPUT} ({interim.Year.nunique()} years)")
    for year in (2026, 2060, 2100):
        year_data = interim[interim.Year == year]
        total = year_data.Total.sum()
        aged_share = year_data[year_data.Age >= 65].Total.sum() / total
        print(f"  {year}: {total / 1e6:.1f}M, 65+ share {aged_share:.1%}")

    manifest = {
        "description": "2026 Trustees Report calibration target sources",
        "retrieved": "2026-06-10",
        "files": [
            {
                "file": str(path.relative_to(REPO_ROOT)),
                "sha256": file_sha256(path),
                "source": SOURCE_URLS.get(path.name, ""),
            }
            for path in (WORKBOOK, MEDICARE_CSV, AUX_OUTPUT, POPULATION_OUTPUT)
        ],
        "notes": [
            "OASDI TOB dollars = IV.B2 percent of taxable payroll times "
            "VI.G1 taxable payroll.",
            "HI TOB from the CMS 2026 Medicare Trustees expanded tables, "
            "annual through 2100 (no carry-forward bridge needed).",
            "Single-year-age population is interim: TR2024 single-year "
            "shape rescaled to TR2026 V.A3 group totals until SSA posts "
            "the TR2026 single-year file.",
            "TR2026 incorporates OBBBA in current law, so these TOB "
            "targets replace the TR2025 post-OBBBA bridge entirely.",
        ],
    }
    MANIFEST_OUTPUT.write_text(json.dumps(manifest, indent=2) + "\n")
    print(f"wrote {MANIFEST_OUTPUT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

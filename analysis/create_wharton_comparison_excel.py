"""
Create Excel spreadsheet with Wharton Budget Model comparison
in the clean table format requested - one sheet with three tables
"""

import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# PolicyEngine and Wharton data
data_2026 = {
    'Income Group': ['First quintile', 'Second quintile', 'Middle quintile', 'Fourth quintile',
                     '80-90%', '90-95%', '95-99%', '99-99.9%', 'Top 0.1%'],
    'PolicyEngine': [-24, -65, -417, -763, -2148, -2907, -1972, -1608, 0],
    'Wharton': [0, -15, -340, -1135, -1625, -1590, -2020, -2205, -2450],
    'Difference': [-24, -50, -77, 372, -523, -1317, 48, 597, 2450],
    '% Diff': ['N/A', '333%', '23%', '-33%', '32%', '83%', '-2%', '-27%', '-100%']
}

data_2034 = {
    'Income Group': ['First quintile', 'Second quintile', 'Middle quintile', 'Fourth quintile',
                     '80-90%', '90-95%', '95-99%', '99-99.9%', 'Top 0.1%'],
    'PolicyEngine': [-39, -195, -769, -1291, -3053, -3388, -2325, -2250, 0],
    'Wharton': [0, -45, -615, -1630, -2160, -2160, -2605, -2715, -2970],
    'Difference': [-39, -150, -154, 339, -893, -1228, 280, 465, 2970],
    '% Diff': ['N/A', '333%', '25%', '-21%', '41%', '57%', '-11%', '-17%', '-100%']
}

data_2054 = {
    'Income Group': ['First quintile', 'Second quintile', 'Middle quintile', 'Fourth quintile',
                     '80-90%', '90-95%', '95-99%', '99-99.9%', 'Top 0.1%'],
    'PolicyEngine': [-5, -242, -757, -1558, -3518, -5094, -5183, -3231, 0],
    'Wharton': [-5, -275, -1730, -3560, -4075, -4385, -4565, -4820, -5080],
    'Difference': [0, 33, 973, 2002, 557, -709, -618, 1589, 5080],
    '% Diff': ['0% ✓', '-12%', '-56%', '-56%', '-14%', '16%', '14%', '-33%', '-100%']
}

data_2054_local = {
    'Income Group': ['First quintile', 'Second quintile', 'Middle quintile', 'Fourth quintile',
                     '80-90%', '90-95%', '95-99%', '99-99.9%', 'Top 0.1%'],
    'PolicyEngine': [-312, -1119, -2982, -4342, -9064, -13974, -6113, -6406, -280],
    'Wharton': [-5, -275, -1730, -3560, -4075, -4385, -4565, -4820, -5080],
    'Difference': [-307, -844, -1252, -782, -4989, -9589, -1548, -1586, 4800],
    '% Diff': ['6240%', '307%', '72%', '22%', '122%', '219%', '34%', '33%', '-94%']
}

data_2054_new = {
    'Income Group': ['First quintile', 'Second quintile', 'Middle quintile', 'Fourth quintile',
                     '80-90%', '90-95%', '95-99%', '99-99.9%', 'Top 0.1%'],
    'PolicyEngine': [-134, -868, -1946, -2644, -4067, -6741, -3097, -4098, -188],
    'Wharton': [-5, -275, -1730, -3560, -4075, -4385, -4565, -4820, -5080],
    'Difference': [-129, -593, -216, 916, 8, -2356, 1468, 722, 4892],
    '% Diff': ['2580%', '216%', '12% ✓', '-26%', '~0% ✓✓', '54%', '-32%', '-15%', '-96%']
}

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Wharton Comparison"

# Define styles
header_font = Font(bold=True, size=14)
table_header_font = Font(bold=True, size=11)
regular_font = Font(size=11)
centered = Alignment(horizontal='center', vertical='center')
left_aligned = Alignment(horizontal='left', vertical='center')
right_aligned = Alignment(horizontal='right', vertical='center')

# Border style
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Fill styles
gray_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

current_row = 1

# Helper function to add a table
def add_table(ws, start_row, year, data, title):
    """Add a formatted table to the worksheet"""
    row = start_row

    # Add title
    ws.cell(row=row, column=1, value=title)
    ws.cell(row=row, column=1).font = header_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1

    # Add header row
    headers = ['Income Group', 'PolicyEngine', 'Wharton', 'Difference', '% Diff']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = table_header_font
        cell.alignment = centered
        cell.fill = header_fill
        cell.border = thin_border
    row += 1

    # Add data rows
    for i, group in enumerate(data['Income Group']):
        # Determine if this row should be gray
        is_gray = i % 2 == 1

        # Income Group
        cell = ws.cell(row=row, column=1, value=group)
        cell.font = regular_font
        cell.alignment = left_aligned
        cell.border = thin_border
        if is_gray:
            cell.fill = gray_fill

        # PolicyEngine
        cell = ws.cell(row=row, column=2, value=f"-${abs(data['PolicyEngine'][i]):,}" if data['PolicyEngine'][i] != 0 else "$0")
        cell.font = regular_font
        cell.alignment = right_aligned
        cell.border = thin_border
        if is_gray:
            cell.fill = gray_fill

        # Wharton
        wh_val = data['Wharton'][i]
        cell = ws.cell(row=row, column=3, value=f"-${abs(wh_val):,}" if wh_val < 0 else f"${wh_val:,}" if wh_val > 0 else "$0")
        cell.font = regular_font
        cell.alignment = right_aligned
        cell.border = thin_border
        if is_gray:
            cell.fill = gray_fill

        # Difference
        diff_val = data['Difference'][i]
        cell = ws.cell(row=row, column=4, value=f"+${abs(diff_val):,}" if diff_val > 0 else f"-${abs(diff_val):,}" if diff_val < 0 else "$0")
        cell.font = regular_font
        cell.alignment = right_aligned
        cell.border = thin_border
        if is_gray:
            cell.fill = gray_fill

        # % Diff
        cell = ws.cell(row=row, column=5, value=data['% Diff'][i])
        cell.font = regular_font
        cell.alignment = centered
        cell.border = thin_border
        if is_gray:
            cell.fill = gray_fill

        row += 1

    return row + 1  # Return next available row with spacing

# Set column widths
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 12

# Add revenue summary at top
ws.cell(row=current_row, column=1, value="AGGREGATE REVENUE IMPACT (Billions)")
ws.cell(row=current_row, column=1).font = header_font
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
current_row += 1

# Revenue data
revenue_data = [
    ['Year 2026:', '-$85.4B', '(Enhanced CPS 2024)'],
    ['Year 2034:', '-$131.7B', '(Enhanced CPS 2024)'],
    ['Year 2054:', '-$176.3B', '(Enhanced CPS 2024)'],
    ['Year 2054 (Old Local):', '-$588.1B', '(2054.h5 - old local)'],
    ['Year 2054 (New):', '-$284.3B', '(2054 (1).h5 - best Wharton match)'],
]
for row_data in revenue_data:
    ws.cell(row=current_row, column=1, value=row_data[0]).font = Font(bold=True, size=11)
    ws.cell(row=current_row, column=2, value=row_data[1]).font = Font(size=11)
    if len(row_data) > 2:
        ws.cell(row=current_row, column=3, value=row_data[2]).font = Font(italic=True, size=10)
    current_row += 1

current_row += 2  # Add spacing

# Add 2026 table
current_row = add_table(ws, current_row, 2026, data_2026, "Average Tax Change per Household (Dollars) - Year 2026")

# Add 2034 table
current_row = add_table(ws, current_row, 2034, data_2034, "Average Tax Change per Household (Dollars) - Year 2034")

# Add 2054 table (Enhanced CPS)
current_row = add_table(ws, current_row, 2054, data_2054, "Average Tax Change per Household (Dollars) - Year 2054 (Enhanced CPS 2024)")

# Add 2054 old local table
current_row = add_table(ws, current_row, 2054, data_2054_local, "Average Tax Change per Household (Dollars) - Year 2054 (Old Local Dataset: 2054.h5)")

# Add 2054 new table
current_row = add_table(ws, current_row, 2054, data_2054_new, "Average Tax Change per Household (Dollars) - Year 2054 (New Dataset: 2054 (1).h5 - BEST MATCH)")

# Add dataset note at bottom
ws.cell(row=current_row, column=1, value="Dataset: Enhanced CPS 2024 (reweighted to target years)")
ws.cell(row=current_row, column=1).font = Font(italic=True, size=10)
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)

# Save workbook
output_file = '../data/wharton_comparison_enhanced_cps_2024.xlsx'
wb.save(output_file)

print(f"✓ Excel file created: {output_file}")
print()
print("Single sheet with formatted tables:")
print("  - Revenue summary (2026, 2034, 2054, 2054 new)")
print("  - 2026 comparison table (formatted)")
print("  - 2034 comparison table (formatted)")
print("  - 2054 comparison table - Enhanced CPS 2024 (formatted)")
print("  - 2054 comparison table - New dataset 2054 (1).h5 (formatted)")
print()
print("Formatting includes:")
print("  - Bold headers with gray background")
print("  - Alternating row colors (white/light gray)")
print("  - Borders on all cells")
print("  - Proper currency formatting")
print("  - Centered/aligned text")
print()
print("✓ Complete!")
